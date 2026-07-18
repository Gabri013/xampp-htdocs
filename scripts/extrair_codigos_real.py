#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import sys
import os
import json
sys.stdout.reconfigure(encoding='utf-8')

try:
    from openpyxl import load_workbook
    print("[OK] openpyxl disponivel\n")
except ImportError:
    print("[WARN] openpyxl nao disponivel, tentando xlrd...")
    try:
        import xlrd
        print("[OK] xlrd disponivel\n")
    except ImportError:
        print("[ERROR] Nenhuma biblioteca disponivel!")
        sys.exit(1)

arquivo = r"C:\Users\gabri\Downloads\CADASTRO PRODUTOS JOTEC - 2019 C.xls"

print("="*70)
print("EXTRAIR CODIGOS REAIS DO EXCEL JOTEC")
print("="*70)
print()

if not os.path.exists(arquivo):
    print(f"[ERROR] Arquivo nao encontrado: {arquivo}\n")
    sys.exit(1)

print(f"Arquivo: {os.path.basename(arquivo)}")
print(f"Tamanho: {os.path.getsize(arquivo) / (1024*1024):.2f} MB\n")

try:
    from openpyxl import load_workbook
    print("Lendo com openpyxl...")
    wb = load_workbook(arquivo, data_only=True)
    abas = wb.sheetnames
    print(f"[OK] Arquivo lido! Total de abas: {len(abas)}\n")

    todos_codigos = []
    dados_por_aba = {}

    for aba_idx, aba_nome in enumerate(abas, 1):
        print(f"[ABA {aba_idx}] {aba_nome}")
        print("-" * 70)

        ws = wb[aba_nome]
        codigos_aba = []

        # Ler headers
        headers = []
        for col in range(1, min(ws.max_column + 1, 10)):
            val = ws.cell(1, col).value
            if val:
                headers.append((col, str(val).strip()))

        print(f"  Colunas encontradas: {len(headers)}")

        # Ler dados (primeiras 100 linhas)
        for row_idx in range(2, min(ws.max_row + 1, 100)):
            linha_valores = {}
            tem_dados = False

            for col, header_nome in headers:
                valor = ws.cell(row_idx, col).value
                if valor:
                    tem_dados = True
                    linha_valores[header_nome] = str(valor).strip()

            if tem_dados:
                # Procurar codigo
                codigo = None
                for col_nome in linha_valores:
                    col_lower = col_nome.lower()
                    if 'cod' in col_lower or 'code' in col_lower:
                        codigo = linha_valores[col_nome]
                        break

                if not codigo and headers:
                    codigo = linha_valores.get(headers[0][1])

                if codigo and codigo not in ['', 'Codigo']:
                    codigos_aba.append(codigo)
                    todos_codigos.append({
                        'codigo': codigo,
                        'aba': aba_nome,
                        'linha': row_idx,
                        'dados': linha_valores
                    })

                    # Amostra
                    if len(codigos_aba) <= 3:
                        print(f"    Linha {row_idx}: {codigo}")

        print(f"  Total de codigos: {len(codigos_aba)}\n")
        dados_por_aba[aba_nome] = codigos_aba

    # Analise
    print("="*70)
    print("ANALISE DOS CODIGOS")
    print("="*70)
    print()

    codigos_unicos = sorted(set([c['codigo'] for c in todos_codigos]))
    print(f"Total de codigos unicos: {len(codigos_unicos)}\n")

    print("PRIMEIROS 30 CODIGOS:")
    for i, cod in enumerate(codigos_unicos[:30], 1):
        print(f"  {i:2d}. {cod}")

    print("\n")

    # Detectar padrao
    print("PADRAO DE CODIGO:")
    print("-" * 70)

    primeiros = codigos_unicos[:5]
    if all(c.isdigit() for c in primeiros):
        print("Tipo: Numerico sequencial")
        print(f"Exemplo: {' -> '.join(primeiros)}\n")
    else:
        print("Tipo: Alfanumerico")
        print(f"Exemplo: {', '.join(primeiros)}\n")

    # Salvar JSON
    output_file = r"C:\xampp\htdocs\scripts\codigos_jotec_reais.json"

    saida = {
        "arquivo": os.path.basename(arquivo),
        "total_codigos": len(codigos_unicos),
        "abas_processadas": list(dados_por_aba.keys()),
        "codigos": codigos_unicos,
        "amostra_primeiros_30": codigos_unicos[:30]
    }

    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(saida, f, ensure_ascii=False, indent=2)

    print(f"[OK] Codigos salvos em: {output_file}\n")

    print("="*70)
    print("EXTRACAO COMPLETA COM SUCESSO!")
    print("="*70)

except Exception as e:
    print(f"[ERROR] {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)
