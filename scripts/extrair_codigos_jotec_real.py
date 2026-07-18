#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
EXTRAIR CÓDIGOS REAIS DO EXCEL JOTEC
Lê o arquivo Excel original e extrai os códigos sequenciais reais da JOTEC
"""

import sys
import os
from pathlib import Path

# Tentar importar openpyxl
try:
    from openpyxl import load_workbook
    print("✅ openpyxl disponível\n")
except ImportError:
    print("⚠️  openpyxl não disponível, tentando xlrd...")
    try:
        import xlrd
        print("✅ xlrd disponível\n")
    except ImportError:
        print("❌ Nenhuma biblioteca disponível!")
        sys.exit(1)

import json
from collections import defaultdict

# Arquivo Excel
arquivo = r"C:\Users\gabri\Downloads\CADASTRO PRODUTOS JOTEC - 2019 C.xls"

print("╔════════════════════════════════════════════════════════════════╗")
print("║  🔍 EXTRAIR CÓDIGOS REAIS DO EXCEL JOTEC                      ║")
print("╚════════════════════════════════════════════════════════════════╝\n")

if not os.path.exists(arquivo):
    print(f"❌ Arquivo não encontrado: {arquivo}\n")
    sys.exit(1)

print(f"📁 Arquivo: {Path(arquivo).name}")
print(f"📊 Tamanho: {os.path.getsize(arquivo) / (1024*1024):.2f} MB\n")

try:
    # Tentar com openpyxl primeiro (mais moderno)
    try:
        print("Tentando com openpyxl...")
        wb = load_workbook(arquivo, data_only=True)
        abas = wb.sheetnames
        print(f"✅ Arquivo lido com sucesso!")
        print(f"   Total de abas: {len(abas)}\n")

        dados_extraidos = {}
        codigos_encontrados = set()
        padrao_codigo = None

        for aba_idx, aba_nome in enumerate(abas, 1):
            print(f"📑 ABA {aba_idx}: {aba_nome}")
            print("═" * 60)

            ws = wb[aba_nome]
            linhas_lidas = 0
            codigos_nesta_aba = []

            # Ler headers
            headers = []
            for col in range(1, ws.max_column + 1):
                val = ws.cell(1, col).value
                if val:
                    headers.append((col, str(val).strip()))

            print(f"   Colunas: {len(headers)}")
            if headers:
                print(f"   Headers: {', '.join([h[1][:20] for h in headers[:5]])}")

            # Ler dados
            for row_idx in range(2, min(ws.max_row + 1, 50)):  # Primeiras 50 linhas
                linha_dados = {}
                tem_dados = False

                for col, header_nome in headers:
                    valor = ws.cell(row_idx, col).value
                    if valor:
                        tem_dados = True
                        linha_dados[header_nome] = str(valor).strip()

                if tem_dados:
                    linhas_lidas += 1

                    # Procurar código em diferentes colunas
                    codigo = None
                    for col_nome in linha_dados:
                        col_lower = col_nome.lower()
                        if 'cod' in col_lower or 'código' in col_lower or 'code' in col_lower:
                            codigo = linha_dados[col_nome]
                            break

                    if not codigo:
                        # Se não achou, tentar primeira coluna
                        if headers:
                            codigo = linha_dados.get(headers[0][1])

                    if codigo and codigo not in ['', 'Código']:
                        codigos_nesta_aba.append({
                            'linha': row_idx,
                            'codigo': codigo,
                            'dados': linha_dados
                        })
                        codigos_encontrados.add(codigo)

                        # Amostra dos primeiros 3 códigos
                        if len(codigos_nesta_aba) <= 3:
                            print(f"   • Linha {row_idx}: {codigo}")

            print(f"   Total de registros: {linhas_lidas}")
            print(f"   Códigos encontrados: {len(codigos_nesta_aba)}\n")

            dados_extraidos[aba_nome] = codigos_nesta_aba

        # Análise dos códigos
        print("╔════════════════════════════════════════════════════════════════╗")
        print("║  📊 ANÁLISE DOS CÓDIGOS ENCONTRADOS                           ║")
        print("╚════════════════════════════════════════════════════════════════╝\n")

        print(f"Total de códigos únicos: {len(codigos_encontrados)}\n")

        # Amostra de códigos
        print("AMOSTRA DOS PRIMEIROS 20 CÓDIGOS:")
        for i, codigo in enumerate(sorted(list(codigos_encontrados))[:20], 1):
            print(f"  {i:2d}. {codigo}")

        print("\n")

        # Detectar padrão
        print("PADRÃO DE CÓDIGO DETECTADO:")
        print("═" * 60)

        primeiros_codigos = sorted(list(codigos_encontrados))[:10]

        # Verificar se são números
        codigos_numericos = [c for c in primeiros_codigos if c.isdigit()]
        if codigos_numericos:
            print(f"✅ Códigos sequenciais numéricos: {codigos_numericos[:3]}")
            padrao = "Numérico sequencial"
        else:
            # Verificar padrão alpanumérico
            print(f"✅ Códigos alpanuméricos: {primeiros_codigos[:3]}")
            padrao = "Alfanumérico"

        print(f"\nTipo de padrão: {padrao}")
        print(f"Total de registros encontrados: {len(codigos_encontrados)}\n")

        # Gerar arquivo JSON com códigos
        output_file = r"C:\xampp\htdocs\scripts\codigos_jotec_extraidos.json"

        saida = {
            "arquivo": Path(arquivo).name,
            "abas_processadas": list(dados_extraidos.keys()),
            "total_codigos": len(codigos_encontrados),
            "padrao": padrao,
            "codigos": sorted(list(codigos_encontrados))
        }

        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(saida, f, ensure_ascii=False, indent=2)

        print(f"✅ Códigos salvos em: {output_file}\n")

        print("╔════════════════════════════════════════════════════════════════╗")
        print("║  ✅ EXTRAÇÃO COMPLETA                                         ║")
        print("╚════════════════════════════════════════════════════════════════╝\n")

    except Exception as e:
        print(f"⚠️  Erro com openpyxl: {e}\n")
        print("Tentando com xlrd (formato XLS antigo)...")

        import xlrd
        wb = xlrd.open_workbook(arquivo)

        print(f"✅ Arquivo lido com xlrd!")
        print(f"   Total de abas: {len(wb.sheet_names())}\n")

        for aba_idx, aba_nome in enumerate(wb.sheet_names()[:5]):  # Primeiras 5 abas
            print(f"📑 ABA: {aba_nome}")
            ws = wb.sheet_by_name(aba_nome)
            print(f"   Linhas: {ws.nrows}, Colunas: {ws.ncols}")

            # Headers
            headers = []
            for col in range(ws.ncols):
                val = ws.cell(0, col).value
                if val:
                    headers.append((col, str(val).strip()))

            print(f"   Headers: {', '.join([h[1][:15] for h in headers[:3]])}")

            # Amostra
            print("   Primeiros dados:")
            for row in range(1, min(4, ws.nrows)):
                print(f"      • {[ws.cell(row, col).value for col in range(min(3, ws.ncols))]}")
            print()

except Exception as e:
    print(f"\n❌ Erro ao processar arquivo: {e}\n")
    import traceback
    traceback.print_exc()
    sys.exit(1)

print("✅ Script concluído!\n")
