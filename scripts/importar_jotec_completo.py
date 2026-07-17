#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
IMPORTAR CADASTRO JOTEC - AUTOMATIZADO

Lê arquivo Excel e importa para banco de dados
"""

import sys
import json
from pathlib import Path

# Tentar diferentes bibliotecas
try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import xlrd
    XLRD_AVAILABLE = True
except ImportError:
    XLRD_AVAILABLE = False

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = True

def ler_arquivo_excel(caminho_arquivo):
    """Lê arquivo Excel e retorna dados"""

    arquivo = Path(caminho_arquivo)

    if not arquivo.exists():
        raise Exception(f"Arquivo não encontrado: {caminho_arquivo}")

    ext = arquivo.suffix.lower()
    dados = {}

    print(f"📂 Arquivo: {arquivo.name}")
    print(f"📊 Tamanho: {arquivo.stat().st_size / 1024 / 1024:.2f} MB\n")

    # Tentar com openpyxl (moderno)
    if ext == '.xlsx' and OPENPYXL_AVAILABLE:
        print("✅ Usando openpyxl para ler .xlsx\n")
        wb = openpyxl.load_workbook(str(arquivo))

        for sheet_name in wb.sheetnames:
            print(f"📑 Aba: {sheet_name}")
            ws = wb[sheet_name]

            # Headers
            headers = []
            for col in range(1, ws.max_column + 1):
                header = ws.cell(1, col).value
                headers.append(header if header else f"Coluna{col}")

            # Dados
            rows = []
            for row_idx in range(2, ws.max_row + 1):
                row_data = {}
                for col_idx, header in enumerate(headers, 1):
                    row_data[header] = ws.cell(row_idx, col_idx).value

                # Pular linhas vazias
                if any(v for v in row_data.values() if v):
                    rows.append(row_data)

            dados[sheet_name] = {
                'headers': headers,
                'rows': rows,
                'total': len(rows)
            }

            print(f"  Linhas: {len(rows)}")
            print(f"  Amostra: {rows[0] if rows else 'Vazio'}\n")

    # Tentar com pandas
    elif PANDAS_AVAILABLE:
        print("✅ Usando pandas para ler arquivo\n")
        xls = pd.ExcelFile(str(arquivo))

        for sheet_name in xls.sheet_names:
            print(f"📑 Aba: {sheet_name}")
            df = pd.read_excel(str(arquivo), sheet_name=sheet_name)

            # Limpar NaN
            df = df.fillna('')

            rows = df.to_dict('records')

            dados[sheet_name] = {
                'headers': list(df.columns),
                'rows': rows,
                'total': len(rows)
            }

            print(f"  Linhas: {len(rows)}")
            if rows:
                print(f"  Amostra: {rows[0]}\n")

    # Tentar com xlrd (antigo)
    elif ext == '.xls' and XLRD_AVAILABLE:
        print("✅ Usando xlrd para ler .xls\n")
        wb = xlrd.open_workbook(str(arquivo))

        for sheet in wb.sheets():
            print(f"📑 Aba: {sheet.name}")

            headers = []
            for col in range(sheet.ncols):
                headers.append(sheet.cell_value(0, col))

            rows = []
            for row_idx in range(1, sheet.nrows):
                row_data = {}
                for col_idx, header in enumerate(headers):
                    row_data[header] = sheet.cell_value(row_idx, col_idx)

                if any(v for v in row_data.values() if str(v).strip()):
                    rows.append(row_data)

            dados[sheet.name] = {
                'headers': headers,
                'rows': rows,
                'total': len(rows)
            }

            print(f"  Linhas: {len(rows)}")
            if rows:
                print(f"  Amostra: {rows[0]}\n")

    else:
        raise Exception(f"Não consegui ler arquivo. Formato: {ext}")

    return dados

def processar_dados(dados):
    """Processa e valida dados"""

    print("\n" + "="*70)
    print("🔍 VALIDANDO DADOS")
    print("="*70 + "\n")

    materias_primas = []
    total_erros = 0

    for aba_nome, aba_dados in dados.items():
        print(f"📑 Aba: {aba_nome} ({aba_dados['total']} linhas)")

        headers = aba_dados['headers']
        rows = aba_dados['rows']

        # Normalizar headers (lowercase, sem espaços)
        headers_norm = {h.lower().strip(): h for h in headers if h}

        # Procurar colunas esperadas
        col_codigo = None
        col_descricao = None
        col_fornecedor = None
        col_preco = None
        col_unidade = None

        for h_norm, h_orig in headers_norm.items():
            if 'cod' in h_norm or 'code' in h_norm:
                col_codigo = h_orig
            elif 'desc' in h_norm or 'name' in h_norm or 'product' in h_norm:
                col_descricao = h_orig
            elif 'forn' in h_norm or 'supplier' in h_norm or 'vendor' in h_norm:
                col_fornecedor = h_orig
            elif 'prec' in h_norm or 'price' in h_norm or 'value' in h_norm:
                col_preco = h_orig
            elif 'unit' in h_norm or 'un' in h_norm:
                col_unidade = h_orig

        print(f"  Colunas detectadas:")
        print(f"    Código: {col_codigo}")
        print(f"    Descrição: {col_descricao}")
        print(f"    Fornecedor: {col_fornecedor}")
        print(f"    Preço: {col_preco}")
        print(f"    Unidade: {col_unidade}\n")

        # Processar linhas
        for idx, row in enumerate(rows, 1):
            try:
                codigo = str(row.get(col_codigo, '')).strip()
                descricao = str(row.get(col_descricao, '')).strip()
                fornecedor = str(row.get(col_fornecedor, '')).strip()
                preco_str = str(row.get(col_preco, '0')).strip()
                unidade = str(row.get(col_unidade, '')).strip()

                # Pular linhas vazias
                if not codigo:
                    continue

                # Validar preço
                try:
                    preco = float(preco_str.replace(',', '.'))
                except:
                    preco = 0.0

                # Validar
                erros = []

                if not codigo:
                    erros.append("Código vazio")
                if not descricao:
                    erros.append("Descrição vazia")
                if preco <= 0:
                    erros.append(f"Preço inválido: {preco_str}")
                if not unidade:
                    erros.append("Unidade vazia")

                if erros:
                    total_erros += 1
                    print(f"  ❌ Linha {idx}: {', '.join(erros)}")
                    continue

                # Adicionar à lista
                materias_primas.append({
                    'aba': aba_nome,
                    'linha': idx,
                    'codigo': codigo,
                    'descricao': descricao,
                    'fornecedor': fornecedor or 'Sem fornecedor',
                    'preco': preco,
                    'unidade': unidade
                })

            except Exception as e:
                total_erros += 1
                print(f"  ❌ Linha {idx}: Erro ao processar - {e}")

        print()

    print(f"✅ Total válido: {len(materias_primas)}")
    print(f"❌ Total com erro: {total_erros}")
    print(f"📊 Taxa de sucesso: {(len(materias_primas) / (len(materias_primas) + total_erros) * 100) if (len(materias_primas) + total_erros) > 0 else 0:.1f}%\n")

    return materias_primas

def salvar_json(dados, arquivo_saida):
    """Salva dados em JSON para processamento PHP"""

    with open(arquivo_saida, 'w', encoding='utf-8') as f:
        json.dump(dados, f, ensure_ascii=False, indent=2)

    print(f"✅ Dados salvos em: {arquivo_saida}")

# Main
if __name__ == '__main__':
    arquivo_entrada = r'C:\Users\gabri\Downloads\CADASTRO PRODUTOS JOTEC - 2019 C.xls'
    arquivo_saida = r'C:\xampp\htdocs\scripts\jotec_dados.json'

    try:
        print("🚀 IMPORTAÇÃO CADAST RO JOTEC")
        print("="*70 + "\n")

        # Ler arquivo
        print("📖 LENDO ARQUIVO EXCEL...\n")
        dados_brutos = ler_arquivo_excel(arquivo_entrada)

        # Processar
        print("\n" + "="*70)
        materias_primas = processar_dados(dados_brutos)

        # Salvar JSON
        print("="*70)
        print("\n💾 SALVANDO DADOS...\n")
        salvar_json(materias_primas, arquivo_saida)

        print(f"\n✅ CONCLUSÃO:")
        print(f"   Arquivo lido: {arquivo_entrada}")
        print(f"   Materiais processados: {len(materias_primas)}")
        print(f"   Próximo passo: Executar importação no banco de dados")

    except Exception as e:
        print(f"\n❌ ERRO: {e}")
        sys.exit(1)
